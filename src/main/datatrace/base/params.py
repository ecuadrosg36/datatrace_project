from openpyxl.styles import PatternFill

WORKSPACE_SHARED_LOG_PATH = '/Workspace/Shared/DataExcellence/Assets/Logs'

## lists
GENERAL_COL_LIST = ["filename_path", "filename", "job_code", "last_modif_date",
                    "q_lines_script", "q_lines_no_coverage", "%_lines_coverage",
                    "q_statements", "q_statements_to_analyze", "q_statements_success_analysis", "%_statement_coverage",
                    "q_comments"]

PREV_BASE_STMT_COL_LIST = ['filename_path', 'filename',
                           'start_line', 'end_line', 'statement_id', 'statement_type', 'statement', 'variable_list',
                           'raw_statement_type', 'raw_statement',
                           'parent_start_line', 'parent_end_line', 'parent_statement_id', 'parent_statement_type',
                           'parent_statement'
                           ]

BASE_STMT_COL_LIST = ['filename_path', 'filename',
                      'start_line', 'end_line', 'statement_id', 'statement_type', 'flag_to_analyze',
                      'flag_mapping_success', 'statement', 'variable_list',
                      'raw_statement_type', 'raw_statement',
                      'parent_start_line', 'parent_end_line', 'parent_statement_id', 'parent_statement_type',
                      'parent_statement'
                      ]

COMMENT_COL_LIST = ['filename', 'job_code', 'last_modif_date', 'comment_id', 'comment']

FIELD_MAPPING_RESUME_LIST = ['filename', 'job_code', 'last_modif_date', 'statement_id', 'statement_type',
                             'input_object_type', 'input_schema_path', 'input_table_name', 'input_subquery_sql', 'input_table_alias', 'input_field',
                             'output_object_type', 'output_schema_path', 'output_table', 'output_subquery_sql', 'output_field_alias',
                             'expression_sql', 'clause', 'statement', 'filter_expression_sql']
PROCEDURES_COL_LIST = ['filename_path', 'filename', 'statement_id', 'statement_type', 'is_control_procedure',
                       'statement']
FIELD_MAPPING_DETAIL_LIST = ['filename', 'statement_id', 'statement_type', 'statement',
                             'flag_get_struct_success', 'stop_step',
                             'nested_query_id', 'nested_query', 'nested_query_name', 'nested_query_refined',
                             'flag_field_mapping_success', 'field_mapping_exception',
                             'input_table_name', 'input_table_alias', 'input_field', 'output_table',
                             'output_field_alias',
                             'expression_sql', 'clause', 'expression_order_clause_level']
EXPORT_COL_LIST = ['process_type', 'job_code', 'level', 'sub_level', 'filename', 'file_path', 'last_modif_date',
                   'source_object_name', 'field_name',
                   'input_object_type', 'input_schema_path', 'input_table_name', 'input_field', 'input_subquery_sql',
                   'output_object_type', 'output_schema_path', 'output_table', 'output_field_alias', 'output_subquery_sql',
                   'expression_sql', 'functional_definition', 'flg_has_impact', 'filter_expression_sql', 'statement', 'cross',
                   'surrogate', 'statement_id', 'statement_type', 'clause'
                   ]

LINEAGE_INPUT_LIST = ['filename', 'input_table', 'input_field', 'clause', 'expression_sql', 'output_table',
                      'output_field']

pattern_list = [r"('''\s*\|\|\s*([^|]*)\s*\|\|\s*''')",  # '''||var||'''
                r"('\s*\|\|\s*([^|]*)\s*\|\|\s*')",  # '||var||'
                r"('\s*\|\|\s*([^|]*))",  # '||var
                r"\|\|\s*('.*?')",  # '||var
                ]
lineage_general_table_trace_list = ['filename', 'source_table', 'target_table', 'level']
lineage_table_origin_destiny_list = ['filename', 'origin_table', 'destiny_table', 'source_table', 'target_table',
                                     'level']
lineage_general_column_trace_list = ['filename', 'source_table', 'source_field', 'target_table', 'target_field',
                                     'level']
lineage_column_origin_destiny_list = ['filename', 'origin_table', 'origin_field', 'destiny_table', 'destiny_field',
                                      'source_table', 'source_field', 'target_table', 'target_field', 'level']

## Dicts
STATEMENT_STRUCT_DICT = {
    "CREATE_TABLE_AS": {
        # "pattern": r"CREATE\s+TABLE\s+(?P<raw_output_table>.+?)\s+(:?TABLESPACE)?\s*(?P<tablespace>.+?)?\s*\bAS\b\s*(?P<raw_select_statement>[^;]+);\s*",
        "pattern": r"CREATE\s+TABLE\s+(?P<raw_output_table>[^\s]+)\s*(?:NOLOGGING\s*)?(?:TABLESPACE\s+(?P<tablespace>[^\s]+)\s*)?\bAS\b\s*(?P<raw_select_statement>[^;]+);",
        "struct_list": ["raw_output_table", "tablespace", "raw_select_statement"]
    },
    "INSERT_INTO_SELECT": {
        "pattern": r"INSERT\s+INTO\s+(?P<raw_output_table>[^\s(]+)(?:\s+NOLOGGING)?\s*(?P<raw_output_table_struct>\([^\)]*\))?\s*(WITH[\s\S]+?)?(SELECT[\s\S]+?);",
        "struct_list": ["raw_output_table", "raw_output_table_struct"]
    },
    "SPOOL_SELECT": {
        # "pattern": r"SPOOL\s+(?P<raw_output_table>[^\s(]+)\s*(?P<raw_select_statement>SELECT[\s\S]+?);",
        "pattern": r"SPOOL\s+(?P<raw_output_table>[^\s;]+)\s*\n(?P<raw_select_statement>(?:\s*SELECT[\s\S]+?)+?)^\s*SPOOL\s+(?:OFF|OUT)\s*;",
        "struct_list": ["raw_output_table", "raw_select_statement"]
    },
    "SPOOL_CUSTOM_SELECT": {
        # "pattern": r"SPOOL\s+(?P<raw_output_table>[^\s(]+)\s*(?P<raw_select_statement>SELECT[\s\S]+?);",
        "pattern": r"SPOOL\s+(?P<raw_output_table>[^\s;]+)\s*(?P<raw_output_table_struct>(PROMPT[^\n]*|\bSELECT\b.+?\bFROM\b\s*DUALs*;))?\s*(?P<raw_select_statement>\bSELECT\b.+?\bFROM\b.+?)\bSPOOL\b\s*(?:OFF|OUT)\s*;",
        "struct_list": ["raw_output_table", "raw_output_table_struct", "raw_select_statement"]
    }
    ,
    "UPDATE": {
        "pattern": r"UPDATE\s+(?P<raw_output_table>.+?)\s+SET\s+(?P<set_clause>.+?)\s+WHERE\s+(?P<where_clause>.+?);",
        "struct_list": ["raw_output_table", "set_clause", "where_clause"]
    },
    "DELETE": {
        "pattern": r"DELETE\s+FROM\s+(?P<raw_output_table>.+?)\s+WHERE\s+(?P<where_clause>.+?);",
        "struct_list": ["raw_output_table", "where_clause"]
    },"BEGIN_END": {
    "pattern" : r"BEGIN\s*([\s\S]+?)\s*END;"
    }
}

GENERAL_STATEMENT_TYPE_DICT = {
    r"WHENEVER\s+[^;]*?;": "WHENEVER",
    r"ALTER\s+SESSION\s+[^;]*?;": "ALTER_SESSION",
    # r"SET\s+[^;]*?;": "SET_PARAMETERS",
    r"SET\s+([a-zA-Z_][a-zA-Z0-9_]*)\s+(.+?);?$": "SET_PARAMETERS",
    r"PROCEDURE\s+\b(?P<proc_name>\w+)\b.*?\s+(?P=proc_name);": "DECLARE_PROCEDURE",
    r"FUNCTION\s+\b(?P<func_name>\w+)\b.*?\s+(?P=func_name);": "DECLARE_FUNCTION",
    # r"DECLARE\s+.*?BEGIN\s+.*?END;": "DECLARE_BEGIN",
    # r"^BEGIN\s+.*?END;": "BEGIN",
    # r"^SPOOL\s+.*?SPOOL\s+OUT\s*;": "SPOOL",
    # r"SPO(?:OL)?\s+(?!&)(?!OFF\b|OUT\b)[^\s;]+\s*\n\s*SELECT.*?^\s*SPO(?:OL)?\s+(?:OFF|OUT)\s*;": "SPOOL_SELECT",
    r"\bSPOOL\b\s+.*?\bSPOOL\b\s*(?:OFF|OUT)\s*;": "SPOOL_SELECT",
    r"SELECT\s+[^;]\bFROM\b[^;]*?;": "SELECT",
    r"BEGIN\s*([\s\S]+?)\s*END;": "BEGIN_END",
    r"INSERT\s+INTO\s+[^;]*\s+SELECT\s+[^;]*?;": "INSERT_INTO_SELECT",
    r"INSERT\s+INTO\s+[^;]*?;": "INSERT_INTO",
    r"UPDATE\s+[^;]*?\bSET\b\s+[^;]*?;": "UPDATE",
    r"DELETE\s+FROM\s+[^;]*?;": "DELETE",
    r"CREATE\s+TABLE\s+[^;]*\s+SELECT\s+[^;]*?;": "CREATE_TABLE_AS",
    r"CREATE\s+TABLE\s+[^;]*?;": "CREATE_TABLE",
    r"ALTER\s+TABLE\s+[^;]*?;": "ALTER_TABLE",
    r"TRUNCATE\s+TABLE\s+[^;]*?;": "TRUNCATE_TABLE",
    r"DROP\s+TABLE\s+[^;]*?;": "DROP_TABLE",
    r"MERGE\s+INTO\s+[^;]*?;": "MERGE_INTO",
    r"EXEC(?:UTE)?\s+IMMEDIATE\s+[^;]*?;": "EXECUTE_IMMEDIATE",
    r"EXEC(?:UTE)?\s+(?!.*PROCEDURE)[^;]*?;": "EXECUTE_PROCEDURE",
    # r"EXEC(?:UTE)?\s+(?!.*(?:IMMEDIATE|_CONTROL))[^;]*?;": "EXECUTE_PROCEDURE",
    r"CURSOR\s+[^;]*?;": "CURSOR",
    r"WHENEVER\s+[^;]*?;": "WHENEVER"
}
CHILD_STATEMENT_TYPE_DICT = {
    r"INSERT\s+INTO\s+[^;]*\s+SELECT\s+[^;]*?;": "INSERT_INTO_SELECT",
    r"INSERT\s+INTO\s+[^;]*?;": "INSERT_INTO",
    r"CREATE\s+TABLE\s+[^;]*\s+SELECT\s+[^;]*?;": "CREATE_TABLE_AS",
    r"DECLARE\s+.*?BEGIN\s+.*?END;": "DECLARE_BEGIN",
    r"BEGIN\s*([\s\S]+?)\s*END;": "BEGIN_END",
    r"CREATE\s+TABLE\s+[^;]*?;": "CREATE_TABLE",
    r"MERGE\s+INTO\s+[^;]*?;": "MERGE_INTO",
    r"UPDATE\s+[^;]*?\bSET\b\s+[^;]*?;": "UPDATE",
    r"DELETE\s+FROM\s+[^;]*?;": "DELETE",
    r"TRUNCATE\s+TABLE\s+[^;]*?;": "TRUNCATE_TABLE",
    r"DROP\s+TABLE\s+[^;]*?;": "DROP_TABLE",
    r"EXEC(?:UTE)?\s+IMMEDIATE\s+[^;]*?;": "EXECUTE_IMMEDIATE",
    r"SELECT\s+[^;]*\s+INTO\s+[^;]*?;": "SELECT_INTO",
    r"SELECT\s+[^;]*?;": "SELECT",
    # r"PROCEDURE\s+.*?IS\s+.*?BEGIN\s+.*?END\s+.*?;": "DECLARE_PROCEDURE",
    # r"FUNCTION\s+.*?IS\s+.*?BEGIN\s+.*?END\s+.*?;": "DECLARE_FUNCTION",
    r"CURSOR\s+[^;]*?;": "CURSOR",
}

COMMENT_TYPE_DICT = {
    r"JOB\s+[:]?\s*(@\w+)": "JOB_CODE",
    r"(\d{2}[/-]\d{2}[/-]\d{4})": "DATE_MODIF"
}

DETAILED_HEADER_CONFIG = {
    "merge_ranges": {
        'PROCESO IMPACTADO': (1, 7),
        'OBJETO FUENTE': (8, 9),
        'ORIGEN DE IMPACTO': (10, 14),
        'DESTINO DEL IMPACTO': (15, 19),
        'DETALLE DEL IMPACTO': (20, 26),
        'DATOS TECNICOS - DATATRACE': (27, 29)
    },
    "fills": {
        'PROCESO IMPACTADO': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        'OBJETO FUENTE': PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"),
        'ORIGEN DE IMPACTO': PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        'DESTINO DEL IMPACTO': PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        'DETALLE DEL IMPACTO': PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid"),
        'DATOS TECNICOS - DATATRACE': PatternFill(start_color="4D93D9", end_color="4D93D9", fill_type="solid")
    },
    "header_level_2": ['Campo Obligatorio'] * 3 + [""] +
                      ['Campo Obligatorio'] * 2 + [""] +
                      ["OBJETO FUENTE", "CAMPO FUENTE"] +
                      ['Campo Obligatorio'] * 4 + ["Campo DataTrace"] +
                      ['Campo Obligatorio'] * 4 + ["Campo DataTrace"] +
                      ['Campo Obligatorio'] * 4 + [""] * 3 +
                      ["Campo DataTrace"] * 3,
    "header_level_3": [
        'TIPO PROCESO', 'JOB', 'NIVEL', 'SUBNIVEL', 'NOMBRE', 'RUTA DEL PROCESO', 'FECHA ULTIMA MODIFICACIÓN',
        'NOMBRE DEL OBJETO FUENTE (ANÁLISIS DE IMPACTO)', 'NOMBRE DEL CAMPO',
        'TIPO OBJETO', 'ESQUEMA/RUTA', 'NOMBRE', 'CAMPO', 'DETALLE SUBCONSULTA',
        'TIPO OBJETO', 'ESQUEMA/RUTA', 'NOMBRE', 'CAMPO', 'DETALLE SUBCONSULTA',
        'CÁLCULOS O TRANSFORMACIONES', 'DEFINICIÓN FUNCIONAL DEL CÁLCULO O TRANSFORMACIONES',
        '¿HAY IMPACTO?', 'DESCRIPCIÓN FILTRO TABLA', 'SCRIPT', 'CROSS', 'SUBROGADA',
        'ID STATEMENT', 'TIPO STATEMENT', 'CLAUSE'
    ]
}