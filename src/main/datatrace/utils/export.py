import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

import datatrace.base.params as params


class ExcelExporter:
    def __init__(self, output_path: str, sheet_config: dict[str, pd.DataFrame]):
        """
        :param output_path: Path where Excel will be saved.
        :param sheet_config: Dict containing sheet names for each section.
        """
        self.output_path = output_path
        self.sheet_config = sheet_config

    def write(self):
        self._export_base_sheets()
        self._append_custom_styled_sheet("datatrace_report", self.sheet_config["datatrace_report"])

    def _export_base_sheets(self):
        """Writes basic DataFrames to their corresponding sheets."""
        with pd.ExcelWriter(self.output_path, engine='xlsxwriter') as writer:
            for sheet_name, df in self.sheet_config.items():
                if sheet_name != "datatrace_report":
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _append_custom_styled_sheet(self, sheet_name: str, df: pd.DataFrame):
        """Appends a new sheet with a custom styled 3-level header."""
        wb = load_workbook(self.output_path)
        ws = wb.create_sheet(title=sheet_name)
        self._write_custom_header_sheet(ws, df)
        wb.save(self.output_path)

    @staticmethod
    def _write_custom_header_sheet(ws: Worksheet, df: pd.DataFrame):
        """Applies a 3-level header and formatting to a new worksheet."""
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold = Font(bold=True)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                             bottom=Side(style="thin"))

        # Header level 1
        for label, (start_col, end_col) in params.DETAILED_HEADER_CONFIG["merge_ranges"].items():
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=label)
            cell.fill = params.DETAILED_HEADER_CONFIG["fills"][label]
            cell.font = bold
            cell.alignment = center
            cell.border = thin_border

        # Header level 2 and 3
        for col_idx, (lvl2, lvl3) in enumerate(
                zip(params.DETAILED_HEADER_CONFIG["header_level_2"], params.DETAILED_HEADER_CONFIG["header_level_3"]), start=1):
            section = next(label for label, (start, end) in params.DETAILED_HEADER_CONFIG["merge_ranges"].items() if
                           start <= col_idx <= end)
            fill = params.DETAILED_HEADER_CONFIG["fills"][section]

            cell2 = ws.cell(row=2, column=col_idx, value=lvl2)
            cell2.fill = fill
            cell2.font = bold
            cell2.alignment = center
            cell2.border = thin_border

            cell3 = ws.cell(row=3, column=col_idx, value=lvl3)
            cell3.fill = fill
            cell3.font = bold
            cell3.alignment = center
            cell3.border = thin_border

            ws.column_dimensions[cell3.column_letter].width = 12

        # Write DataFrame
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=4):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)