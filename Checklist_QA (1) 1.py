# Databricks notebook source
# coding=utf-8 
# ||********************************************************************************************************
# || PROYECTO               : DATA OPS - QA_VENDOR - CHECKLIST
# || NOMBRE                 : QA_MODELER_CHECKLIST.py
# || TABLA DESTINO          : NA
# || TABLAS FUENTES         : NA
# || OBJETIVO               : Creación de CHECKLIST
# || TIPO                   : pyspark
# || REPROCESABLE           : SI
# || SCHEDULER              : NA
# || JOB                    : NA
# || VERSION    DESARROLLADOR           PROVEEDOR        PO               FECHA             DESCRIPCION
# || 1        ROXANA BECERRA LOPEZ       BCP       URBANO QUISPE     	2025-05-27       Creación del checklist
# *******************************************************************************************************************************

# COMMAND ----------

# Librerias

import pandas as pd
import subprocess, os, unicodedata, re
import unicodedata
from openpyxl import load_workbook
# EXCLUDE_ANALYZER: DataFrame.collect
from pyspark.sql.functions import count, when, col, trim, lower, lit, collect_set, size, array_contains, upper, translate, row_number, monotonically_increasing_id, length, substring
from functools import reduce
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from pyspark.sql import DataFrame, functions as F
from pyspark.sql.window import Window
from typing import List
from openpyxl.styles.protection import Protection
from pyspark.sql.types import StringType
from datetime import datetime


# COMMAND ----------

dbutils.widgets.text("aa_user", "micorreo@bcp.com.pe")
aa_user = dbutils.widgets.get("aa_user")
dbutils.widgets.text("ab_carpeta", "mi_carpeta")
ab_carpeta = dbutils.widgets.get("ab_carpeta").lower()
base_path = f"/Workspace/Users/{aa_user}/"
base_path_carpeta = f"/Workspace/Users/{aa_user}/{ab_carpeta}/"

# COMMAND ----------


### validando que exista la carpeta indicada en el parametro:
# Busca la carpeta ignorando mayusculas
result = subprocess.run(["ls", base_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
folders_list = result.stdout.splitlines()

match = next((f for f in folders_list if f.lower() == ab_carpeta), None)

if match:
    base_path_carpeta = f"{base_path}{match}/"
else:
    raise ValueError(f"En la ruta:  {base_path}  no se encontró la carpeta: {ab_carpeta} especificada en el parametro")


# COMMAND ----------

###Definimos las funciones necesarias
#Funcion para leer los archivos excel
def read_excel_files(
    ruta: str,
    prefijo: str or list = None,
    extension: str = None,
    hoja: str = "Sheet1",
    skip_filas: int = 0,
    agregar_nombre_archivo: bool = False,
    archivo_especifico: str = None
) -> DataFrame:
    """
    Lee uno o varios archivos Excel de una carpeta, según prefijo, extensión y hoja específica.
 
    Args:
        ruta (str): Carpeta donde están los archivos.
        prefijo (str or list, opcional): Prefijo o lista de prefijos que deben tener los archivos.
        extension (str, opcional): Extensión de archivo, como ".xlsx" o ".xlsm".
        hoja (str, opcional): Nombre de la hoja de Excel a leer.
        skip_filas (int, opcional): Número de filas a omitir.
        agregar_nombre_archivo (bool, opcional): Si se agrega columna con nombre de archivo.
        archivo_especifico (str, opcional): Si se indica, se leerá solo ese archivo.
 
    Returns:
        DataFrame: DataFrame de Spark con los datos consolidados.
    """
 
    if archivo_especifico:
        archivos_filtrados = [os.path.join(ruta, archivo_especifico)]
    else:
        if isinstance(prefijo, list):
            prefijo = tuple(prefijo)  # convierte lista a tupla para usar en startswith
 
        archivos = [
            f for f in os.listdir(ruta)
            if (prefijo is None or f.startswith(prefijo))
            and (extension is None or f.endswith(extension))
        ]
        archivos_filtrados = [os.path.join(ruta, f) for f in archivos]
 
    if not archivos_filtrados:
        raise FileNotFoundError(f"No se encontraron archivos con el prefijo:{prefijo} proporcionadas en {ruta}")
 
    df_list = []
    for archivo in archivos_filtrados:
        try:
            xls = pd.ExcelFile(archivo)
            if hoja in xls.sheet_names:
                df = pd.read_excel(archivo, sheet_name=hoja, skiprows=skip_filas, engine="openpyxl")
                if agregar_nombre_archivo:
                    df["ARCHIVO"] = os.path.basename(archivo)
 
                if not df.empty:
                    df_list.append(df)
                else:
                    print(f" Archivo {archivo} está vacío, omitiéndolo.")
            else:
                print(f" El archivo {archivo} no tiene la hoja '{hoja}'.")
        except Exception as e:
            print(f" Error al leer {archivo}: {e}")
 
    if df_list:
        df_consolidado_pd = pd.concat(df_list, ignore_index=True).dropna(how="all")
        df_consolidado_pd.columns = [col.upper() for col in df_consolidado_pd.columns]
        return spark.createDataFrame(df_consolidado_pd)
    else:
        raise ValueError("No se pudieron consolidar los archivos correctamente. Verifica que los archivos sean válidos y contengan datos en la hoja especificada.")

# Funcion para eliminar las tildes de los nombres de campo input df:
def remove_accents_df(df: DataFrame) -> DataFrame:
    def remove_accents(text: str) -> str:
        return ''.join(c for c in unicodedata.normalize('NFD', text) if not unicodedata.combining(c))
    new_columns = [remove_accents(col) for col in df.columns]
    return df.toDF(*new_columns)

#Funcion para eliminar las tildes en el texto:
def remove_accents(text):
    if text is None:
        return ""
    return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn') #Mark,Nonspacing


#Obtener campos
def read_fields(df, capa, validar=None):
    # Generamos una columna ORDEN usando monotonically_increasing_id()
    df = df.withColumn("ORDEN", monotonically_increasing_id())

    """
    Lee los registros del archivo Controles_Parent_domain y los carga como lista.
 
    Args:
        df: En este caso el df del control_parent_domain.
        capa: Capa a validar.
    Returns:
        list_fields: Lista de campos a mostrar.
        nom_prefijos: Nombre de los prefijos de los archivos a considerar.
        nom_hojas: Nombre de las hojas de los archivos a considerar.
        
    """
 
    # Filtrado por capa y validar
    df_filter = df.filter(upper(col("CAPA")) == capa.upper())
    if validar is not None:
        df_filter = df_filter.filter(upper(col("VALIDAR")) == validar.upper())
 
    # Función auxiliar para extraer valores únicos conservando orden
    def extract_ordered_unique_values(df_interno, col_name):
        window_spec = Window.partitionBy(col_name).orderBy("ORDEN")
        df_ranked = df_interno.withColumn("rn", row_number().over(window_spec))
        result_rows = (
            df_ranked.filter(col("rn") == 1)
            .select(col_name, "ORDEN")
            .where(col(col_name).isNotNull())
            .orderBy("ORDEN")
            .collect()
        )
        return [remove_accents(row[col_name].upper()) for row in result_rows]
 
    # NOM_CAMPO_ARCHIVO
    if "NOM_CAMPO_ARCHIVO" in df_filter.columns:
        list_fields = extract_ordered_unique_values(df_filter, "NOM_CAMPO_ARCHIVO")
    else:
        raise ValueError("No se encontró el campo: NOM_CAMPO_ARCHIVO")
 
    # NOM_PREFIJO_ARCHIVO
    if "NOM_PREFIJO_ARCHIVO" in df_filter.columns:
        nom_prefijos = extract_ordered_unique_values(df_filter, "NOM_PREFIJO_ARCHIVO")
    else:
        raise ValueError("No se encontró el campo: NOM_PREFIJO_ARCHIVO")
 
    # NOM_HOJA
    if "NOM_HOJA" in df_filter.columns:
        nom_hoja = extract_ordered_unique_values(df_filter, "NOM_HOJA")
    else:
        raise ValueError("No se encontró el campo: NOM_HOJA")
 
    return list_fields, nom_prefijos, nom_hoja


# Seleccionamos las columnas requeridas y validamos q existan en los archivos DA
def validate_columns(df: DataFrame, required_columns: List[str], df_name: str) -> None:
    missing = set(required_columns) - set(df.columns)
    if missing:
        raise ValueError(f"El archivo '{df_name}' no contiene las siguientes columnas necesarias para la validación: {missing}")

# Funcion para eliminar valores nulos
def clear_column(df, c):
    return df.withColumn(c, when(trim(lower(col(c))).isin("nan", ".."), "").otherwise(trim(col(c)))
    )

# Normalizar columnas
def normalize_and_names(df: DataFrame, to_upper_columns: list) -> DataFrame:
    new_columns_names = [c.strip().replace(" ","_") for c in df.columns]
    df = df.toDF(*new_columns_names)

    for c in to_upper_columns:
        c_norm = c.strip().replace(" ","_")
        if c_norm in df.columns:
            df = df.withColumn(c_norm, upper(trim(col(c_norm))))
        else:
            print(f" Columna '{c}' no existe en el dataframe")
    
    return df

# Validar campos DA vs QA
def process_dataframes(da_df: DataFrame, qa_df: DataFrame, tabla: str, campo: str) -> DataFrame:
    #Guardamos el orden de los campos del df
    columns_order = da_df.columns

 
    qa_comp = qa_df.selectExpr(
            f"TABLA_FISICA as {tabla}",
            f"CAMPO_FISICO as {campo}",
            "True as FLAG_QA"
        )

    #DA con flag
    da_df_flag = da_df.selectExpr("*", "True as FLAG_DA")
    #Join
    joined_df = da_df_flag.join(
        qa_comp,
        on=[tabla, campo],
        how="full"
    )
    #QA_Campo con mensaje segun caso
    resultado = joined_df.withColumn(
        "QA_Campo",
        when(col("FLAG_DA").isNotNull() & col("FLAG_QA").isNotNull(), "OK")
        .when(col("FLAG_DA").isNotNull() & col("FLAG_QA").isNull(), "No esta modelado")
        .when(col("FLAG_DA").isNull() & col("FLAG_QA").isNotNull(), "En QA pero no en DA")
    ).drop("FLAG_DA","FLAG_QA")

    total = resultado.count()
    faltantes = resultado.filter(col("QA_Campo") != "OK").count()
    if faltantes == 0:
        print(f"✅ Todos los {total} campos de {capa['nombre']} coinciden con QA.")
    else:
        print(f"⚠️ De {total} campos de {capa['nombre']}, {faltantes} no están en QA.")
    
    #Reordenamos en base a la lista de columnas del df original
    columns_end = [c for c in columns_order if c in resultado.columns] + \
                  [c for c in resultado.columns if c not in columns_order]
    return resultado.select(*columns_end)

# Funcion para consolidar los resultados de cada capa:
def standardize_columns(df, sufijo):
    """
    Renombra las columnas de un df quitando el sufijo(como: DDV o UDV) para unirlas
    """
    columns_rename = []
    for c in df.columns:
        #Quitamos sufijo de las columnas
        new_name = re.sub(rf"\s*{sufijo}$", "", c.strip(), flags = re.IGNORECASE)
        columns_rename.append(col(c).alias(new_name))
    return df.select(columns_rename)


# COMMAND ----------

###Parametros de validacion###
#Leemos el archivo Controles_Parent_domain
df_control = read_excel_files(
    ruta=base_path,
    prefijo="Controles_Parent_domain",
    extension=".xlsx",
    hoja="parametros",
    skip_filas=0,
    agregar_nombre_archivo=False
)
#Obtenemos los parametros DDV
fields_ddv, prefijo_ddv, hoja_ddv = read_fields(df_control, "DDV")
# #Obtenemos los campos para validacion
fields_ddv_validation, prefijo_ddv, hoja_ddv = read_fields(df_control, "DDV", "SI")

print(f"Campos DDV:, {fields_ddv}\nPrefijo DDV: {prefijo_ddv}\nHoja DDV: {hoja_ddv}\nCampos DDV para validacion: {fields_ddv_validation}")


#Obtenemos los parametros UDV 
fields_udv, prefijo_udv, hoja_udv = read_fields(df_control, "UDV")
# #Obtenemos los campos para validacion
fields_udv_validation, prefijo_udv, hoja_udv = read_fields(df_control, "UDV", "SI")

#Obtenemos los campos QA
fields_qa, prefijo_qa, hoja_qa = read_fields(df_control, "QA")


# COMMAND ----------

### Leemos los archivos del workspace###
#  Listamos todos los archivos en el directorio incluido su path
output = subprocess.run(["ls", base_path_carpeta], capture_output=True, text=True)
# Capturar los nombres
file_list = output.stdout.split("\n")

# Filtrar los archivos que empiezan con "QA_REPORTE_" y terminan en ".xlsx"
qa_file_list = [base_path_carpeta + f for f in file_list if f.startswith(prefijo_qa[0]) and f.endswith('.xlsx')]
# Filtrar los archivos que empiezan con "Documento_Alcance_" y terminan en ".xlsx"
da_ddv_file_list = [base_path_carpeta + f for f in file_list if f.startswith(prefijo_ddv[0]) and f.endswith('.xlsm')]
# Filtrar los archivos que empiezan con "Documento_Alcance_" y terminan en ".xlsx"
da_udv_file_list = [base_path_carpeta + f for f in file_list if f.startswith(prefijo_udv[0]) and f.endswith('.xlsm')]

# Extraer solo los nombres
qa_file_names = [os.path.basename(f) for f in qa_file_list]
da_ddv_file_names = [os.path.basename(f) for f in da_ddv_file_list]
da_udv_file_names = [os.path.basename(f) for f in da_udv_file_list]
print(f"Nombres de archivos QA_: {qa_file_names}\nNombres de archivos DA_DDV: {da_ddv_file_names}\nNombres de archivos DA_UDV: {da_udv_file_names}")

# COMMAND ----------

### Consolidado del archivo QA_REPORTE_
# 1. Leemos el archivo
df_qa_consolidado = read_excel_files(
    ruta= base_path_carpeta,
    prefijo= prefijo_qa[0],
    extension = ".xlsx",
    hoja = hoja_qa[0],
    skip_filas = 0,
    agregar_nombre_archivo=True
)

# 2. Seleccionamos las columnas requeridas y validamos q existan
validate_columns(df_qa_consolidado, fields_qa, "QA_REPORTE") 

# 3. Funcion para eliminar valores nulos
columns_to_clean = df_qa_consolidado.select(fields_qa + ["ARCHIVO"])
df_qa_consolidado_cleaned = reduce(clear_column, columns_to_clean.columns, columns_to_clean)

# 4. Normalizar columnas
column_qa_norm = df_qa_consolidado_cleaned.drop("ARCHIVO", "PARENT_DOMAIN")
df_qa_consolidado_norm = normalize_and_names(df_qa_consolidado_cleaned,column_qa_norm.columns)

# COMMAND ----------

### Definicion de parametros, limpieza, normalizacion y validacion de campos por cada capa
capas = [
    {
        "nombre": "DDV",
        "prefijo": prefijo_ddv[0],
        "hoja": hoja_ddv[0],
        "campos": fields_ddv,
        "validacion": fields_ddv_validation
    },
    {
        "nombre": "UDV",
        "prefijo": prefijo_udv[0],
        "hoja": hoja_udv[0],
        "campos": fields_udv,
        "validacion": fields_udv_validation
    }
]
 
for capa in capas:
    archivos_capa = [f for f in os.listdir(base_path_carpeta) if f.startswith(capa["prefijo"]) and f.endswith(".xlsm")]
 
    if not archivos_capa:
        print(f"⚠️ No se encontraron archivos con prefijo '{capa['prefijo']}' para capa {capa['nombre']}.")
        continue  # Pasa a la siguiente capa
 
    try:
        df_consolidado = read_excel_files(
            ruta=base_path_carpeta,
            prefijo=capa["prefijo"],
            extension=".xlsm",
            hoja=capa["hoja"],
            skip_filas=5,
            agregar_nombre_archivo=True
        )
        
        campo_col = f'"CAMPO {capa["nombre"]}"'
        df_consolidado = df_consolidado.selectExpr("*", f"row_number() over (ORDER BY {campo_col}) as ORDEN")
        
        # 1. Limpiar tildes
        df_consolidado = remove_accents_df(df_consolidado)

        # 2. Validar columnas 
        validate_columns(df_consolidado, capa["campos"], f"Documento de Alcance {capa['nombre']}")       
        # 3. Limpiar nulos
        columnas_con_archivo = df_consolidado.select(*capa["campos"] + ["ARCHIVO", "ORDEN"])
        df_limpio = reduce(clear_column, columnas_con_archivo.columns, columnas_con_archivo)
        
        # 4. Filtrar filas no vacías
        condition = reduce(lambda a, b: a | b, [(col(c).isNotNull()) & (col(c) != "") for c in capa["campos"]])
        df_filtrado = df_limpio.filter(condition)
               
        # 5. Normalizar nombres de columnas
        df_normalizado = normalize_and_names(df_filtrado, capa["validacion"])
        
        # 6. Comparar contra QA
        tabla_col = f"TABLA_{capa['nombre']}"
        campo_col = f"CAMPO_{capa['nombre']}"
 
        resultado_df = process_dataframes(df_normalizado, df_qa_consolidado_norm, tabla_col, campo_col)
        
        #Obtenemos los campos menos ARCHIVO para ordenarlos:
        columns = [col for col in resultado_df.columns if col not in ["ARCHIVO","ORDEN"]]

        if capa["nombre"] == "DDV":         
            #ponemos primero ARCHIVO para mostrar en el resultado exportado
            df_ddv_resultado = resultado_df.select(["ORDEN","ARCHIVO"] + columns)
           
        else:
            df_udv_resultado = resultado_df.select(["ORDEN","ARCHIVO"] + columns)          
 
    except FileNotFoundError as e:
        print(f"❌ Error en lectura de archivos para {capa['nombre']}: {e}")
        


# COMMAND ----------

# Validamos que existan uno o ambos DA
try:
    df_ddv_resultado
except NameError:
    df_ddv_resultado = None
 
try:
    df_udv_resultado
except NameError:
    df_udv_resultado = None
 
#Estandarizando
if df_ddv_resultado is not None:
    df_ddv_resultado = standardize_columns(df_ddv_resultado, "DDV")

if df_udv_resultado is not None:
    df_udv_resultado = standardize_columns(df_udv_resultado, "UDV")

# Unión condicional
if df_ddv_resultado is not None and df_udv_resultado is not None:
    df_resultado_da = df_ddv_resultado.unionByName(df_udv_resultado)
elif df_ddv_resultado is not None:
    df_resultado_da = df_ddv_resultado
elif df_udv_resultado is not None:
    df_resultado_da = df_udv_resultado
else:
    print("No se encontraron resultados")

# COMMAND ----------

###Parametros de validacion###
#Leemos el archivo de controles
df_control_parent_domain = read_excel_files(
    ruta=base_path,
    prefijo="Controles_Parent_domain",
    extension=".xlsx",
    hoja="controles",
    skip_filas=0,
    agregar_nombre_archivo=False
)

# 1. Funcion para eliminar valores nulos
df_control_parent_domain_unique = df_control_parent_domain.dropDuplicates()
df_control_parent_domain_norm = reduce(clear_column, df_control_parent_domain_unique.columns, df_control_parent_domain_unique) 

# Paso 2: Lineamiento por tipo (Fijo o PreFijo)
comparacion_lineamiento_df = df_qa_consolidado_norm.join(
    df_control_parent_domain_norm.filter(col("LINEAMIENTO") == "Fijo"),
    on = col("CAMPO_FISICO") == col("CONTROL"),
    how = "left"
).select(
    df_qa_consolidado_norm["*"],
    when(col("LINEAMIENTO").isNotNull(),col("LINEAMIENTO")).otherwise("PreFijo").alias("LINEAMIENTO_PD")
)

# Paso 3: Validacion por su lineamiento: Fijo / Prefijo
df_fijo = comparacion_lineamiento_df.filter(col("LINEAMIENTO_PD") == "Fijo")
df_prefijo = comparacion_lineamiento_df.filter(col("LINEAMIENTO_PD") == "PreFijo")

#Para Fijo
if df_fijo.count() > 0:
    comparacion_control_fijo_df = df_fijo.join(
        df_control_parent_domain_norm.select("CONTROL", "LINEAMIENTO"),
        (df_fijo.CAMPO_FISICO == df_control_parent_domain_norm.CONTROL) &
        (df_fijo.LINEAMIENTO_PD == df_control_parent_domain_norm.LINEAMIENTO),
        "left"
    ).select(
        df_fijo["*"],
        df_control_parent_domain_norm.CONTROL.alias("CONTROL_PR"))

df_comparacion_parent_domain_fijo = comparacion_control_fijo_df.join(
    df_control_parent_domain_norm.filter(col("LINEAMIENTO") == "Fijo"),
    (col("CONTROL_PR") == col("CONTROL")),
    "left"
).select(
    comparacion_control_fijo_df["*"],
    when (comparacion_control_fijo_df.PARENT_DOMAIN == df_control_parent_domain_norm.PARENT_DOMAIN, "OK")\
        .otherwise(col("WARNING_PARENT_DOMAIN")).alias("Observaciones")).drop("LINEAMIENTO_PD","CONTROL_PR")
 


#Lista de los controles PreFijo
prefijo_fields_len = (
    df_control_parent_domain_norm
    .filter(col("LINEAMIENTO") == "PreFijo")
    .select(
        col("CONTROL"),
        length(trim(col("CONTROL"))).cast("int").alias("LEN_CONTROL"),
        col("PARENT_DOMAIN")
    )
    .dropDuplicates()
    .collect()
)

# Ordenamos por longitud descendente para priorizar los más específicos
prefijo_fields_len = sorted(prefijo_fields_len, key=lambda row: row["LEN_CONTROL"], reverse=True)

# list_prefijo_fields = [row["CONTROL"].upper() for row in prefijo_fields_len]
#  
#Para PreFijo
if df_prefijo.count() > 0:
    #Creamos una columna CONTROL_PR  basada en que si comienza con algun prefijo
    expr = F.lit("OBSERVADO")
    for row in reversed(prefijo_fields_len):  
        prefix = row["CONTROL"].upper()
        length_prefix = row["LEN_CONTROL"]

        expr = when(
            substring(trim(upper(col("CAMPO_FISICO"))), 1, length_prefix) == prefix,
            prefix
        ).otherwise(expr)


    comparacion_control_prefijo_df = df_prefijo.select(
        "*",
        expr.alias("CONTROL_PR")
    )

df_comparacion_parent_domain_prefijo = comparacion_control_prefijo_df.join(
    df_control_parent_domain_norm.filter(col("LINEAMIENTO") == "PreFijo"),
    (col("CONTROL_PR") == col("CONTROL")),
    "left"
).select(
    comparacion_control_prefijo_df["*"],
    when (comparacion_control_prefijo_df.PARENT_DOMAIN == df_control_parent_domain_norm.PARENT_DOMAIN, "OK")\
        .otherwise(col("WARNING_PARENT_DOMAIN")).alias("Observaciones")).drop("LINEAMIENTO_PD","CONTROL_PR")

# Paso 4: Unimos los resulatdos de los controles Fijos y Prefijos
if df_fijo.count() > 0 and df_prefijo.count() > 0:
    comparacion_control_df = df_comparacion_parent_domain_fijo.unionByName(df_comparacion_parent_domain_prefijo)
elif df_fijo.count() > 0:
    comparacion_control_df = df_comparacion_parent_domain_fijo
elif df_prefijo.count() > 0:
    comparacion_control_df = df_comparacion_parent_domain_prefijo
else:
    comparacion_control_df = None

# Paso 5: Validación de Parent Domain
columns = [col for col in comparacion_control_df.columns if col != "ARCHIVO"]
df_comparacion_parent_domain = comparacion_control_df.select(["ARCHIVO"] + columns)
df_comparacion_parent_domain = df_comparacion_parent_domain \
    .filter(col("ORDEN_FISICO").isNotNull()) \
    .selectExpr("*", "cast(ORDEN_FISICO as int) as ORDEN_FISICO_2") \
    .orderBy("TABLA_FISICA", col("ORDEN_FISICO").cast("int"))


# COMMAND ----------

# Obtener la fecha y hora actual
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
# Convertir los DataFrames de Spark a Pandas
# EXCLUDE_ANALYZER: DataFrame.toPandas
pd_resultado_da = df_resultado_da.toPandas()
pd_resultado_da["ORDEN"] = pd.to_numeric(pd_resultado_da["ORDEN"], errors="coerce")
pd_resultado_da = pd_resultado_da.sort_values(by="ORDEN", ascending=True)
# EXCLUDE_ANALYZER: DataFrame.toPandas
resultado_parent_domain_pd = df_comparacion_parent_domain.toPandas()
resultado_parent_domain_pd = resultado_parent_domain_pd.drop("ORDEN_FISICO_2", axis=1)
# Ruta del archivo Excel con ambas hojas
path_export = f"{base_path_carpeta}Reporte_Validacion_QA_{ab_carpeta}_{timestamp}.xlsx"
 
# Colores y estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
zebra_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB34", end_color="FFEB34", fill_type="solid")
alert_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja alerta
white_font = Font(color="FFFFFF", bold=True)
 
# Crear Excel con múltiples hojas
with pd.ExcelWriter(path_export, engine="openpyxl") as writer:
 
    # ------------------ Hoja Consolidado_DA ------------------ #
    pd_resultado_da.to_excel(writer, sheet_name="Consolidado_DA", index=False)
    ws1 = writer.sheets["Consolidado_DA"]
 
    num_cols_1 = len(pd_resultado_da.columns)
    num_rows_1 = len(pd_resultado_da)
    estado_col_index_1 = pd_resultado_da.columns.get_loc("QA_Campo") + 1
 
    for col in range(1, num_cols_1 + 1):
        cell = ws1.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = white_font
 
    for row in range(2, num_rows_1 + 2):
        estado_cell = ws1.cell(row=row, column=estado_col_index_1)
        apply_yellow = estado_cell.value != "OK"
        apply_alert = estado_cell.value == "En QA pero no en DA"
        for col in range(1, num_cols_1 + 1):
            cell = ws1.cell(row=row, column=col)
            if apply_alert:
                cell.fill = alert_fill
            elif apply_yellow:
                cell.fill = yellow_fill
            elif row % 2 == 0:
                cell.fill = zebra_fill
 
    for col_idx, column in enumerate(pd_resultado_da.columns, start=1):
        max_length = max(pd_resultado_da[column].astype(str).map(len).max(), len(column))
        ws1.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
 
    # ------------------ Hoja QA ------------------ #
    resultado_parent_domain_pd.to_excel(writer, sheet_name="QA_MODELO", index=False)
    ws2 = writer.sheets["QA_MODELO"]
 
    num_cols_2 = len(resultado_parent_domain_pd.columns)
    num_rows_2 = len(resultado_parent_domain_pd)
    estado_col_index_2 = resultado_parent_domain_pd.columns.get_loc("Observaciones") + 1
 
    for col in range(1, num_cols_2 + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = white_font
 
    for row in range(2, num_rows_2 + 2):
        estado_cell = ws2.cell(row=row, column=estado_col_index_2)
        apply_yellow = estado_cell.value != "OK"
        for col in range(1, num_cols_2 + 1):
            cell = ws2.cell(row=row, column=col)
            if apply_yellow:
                cell.fill = yellow_fill
            elif row % 2 == 0:
                cell.fill = zebra_fill
 
    for col_idx, column in enumerate(resultado_parent_domain_pd.columns, start=1):
        max_length = max(resultado_parent_domain_pd[column].astype(str).apply(len).max(), len(column))
        ws2.column_dimensions[get_column_letter(col_idx)].width = max_length + 2