# Databricks notebook source
# coding=utf-8 
# ||********************************************************************************************************
# || PROYECTO               : DATA OPS - QA_VENDOR - CHECKLIST
# || NOMBRE                 : QA_MODELER_CHECKLIST_MEJORADO.py
# || TABLA DESTINO          : NA
# || TABLAS FUENTES         : NA
# || OBJETIVO               : Creación de CHECKLIST con mejoras de logging, validación y métricas
# || TIPO                   : pyspark
# || REPROCESABLE           : SI
# || SCHEDULER              : NA
# || JOB                    : NA
# || VERSION    DESARROLLADOR           PROVEEDOR        PO               FECHA             DESCRIPCION
# || 1        ROXANA BECERRA LOPEZ       BCP       URBANO QUISPE     	2025-05-27       Creación del checklist
# || 2        ENMANUEL CUADROS           BCP       URBANO QUISPE        2025-10-21       Optimización y mejoras
# *******************************************************************************************************************************

# COMMAND ----------

# ========================================
# 🆕 NUEVO: Importaciones organizadas por categoría
# ========================================

# Librerías estándar
import pandas as pd
import subprocess
import os
import unicodedata
import re
import shutil
import logging
from datetime import datetime
from typing import List, Tuple, Optional
from functools import reduce

# Librerías de Excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.protection import Protection

# PySpark
# EXCLUDE_ANALYZER: DataFrame.collect
from pyspark.sql import DataFrame, functions as F
from pyspark.sql.functions import (
    count, when, col, trim, lower, lit, collect_set, size, 
    array_contains, upper, translate, row_number, 
    monotonically_increasing_id, length, substring
)
from pyspark.sql.window import Window
from pyspark.sql.types import StringType

# COMMAND ----------

# ========================================
# 🆕 NUEVO: Sistema de Logging Estructurado
# ========================================

def setup_logger(base_path_carpeta: str, ab_carpeta: str) -> logging.Logger:
    """
    Configura el sistema de logging con archivo y consola.
    
    Args:
        base_path_carpeta: Ruta base de la carpeta
        ab_carpeta: Nombre de la carpeta
        
    Returns:
        Logger configurado
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Crear carpeta de logs si no existe
    log_dir = f"{base_path_carpeta}logs/"
    os.makedirs(log_dir, exist_ok=True)
    
    log_file = f"{log_dir}qa_checklist_{ab_carpeta}_{timestamp}.log"
    
    # Configuración del logger
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - [%(funcName)s] - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.info("="*80)
    logger.info(f"🚀 Iniciando proceso de validación QA para carpeta: {ab_carpeta}")
    logger.info("="*80)
    
    return logger

# COMMAND ----------

# ========================================
# PARÁMETROS DE ENTRADA
# ========================================

dbutils.widgets.text("aa_user", "micorreo@bcp.com.pe")
aa_user = dbutils.widgets.get("aa_user")

dbutils.widgets.text("ab_carpeta", "mi_carpeta")
ab_carpeta = dbutils.widgets.get("ab_carpeta").lower()

# 🆕 NUEVO: Widget para modo dry-run (testing sin exportar)
dbutils.widgets.dropdown("dry_run", "No", ["Si", "No"])
dry_run = dbutils.widgets.get("dry_run") == "Si"

# 🆕 NUEVO: Widget para número máximo de versiones a mantener
dbutils.widgets.text("max_versiones", "10")
max_versiones = int(dbutils.widgets.get("max_versiones"))

base_path = f"/Workspace/Users/{aa_user}/"
base_path_carpeta = f"/Workspace/Users/{aa_user}/{ab_carpeta}/"

# COMMAND ----------

# ========================================
# 🆕 NUEVO: Validación de Entrada Robusta
# ========================================

def validate_inputs(aa_user: str, base_path: str, ab_carpeta: str) -> None:
    """
    Valida parámetros de entrada antes de iniciar el proceso.
    
    Args:
        aa_user: Email del usuario
        base_path: Ruta base
        ab_carpeta: Nombre de carpeta
        
    Raises:
        ValueError: Si alguna validación falla
    """
    errors = []
    
    # Validar formato de email
    if not re.match(r"[^@]+@bcp\.com\.pe", aa_user):
        errors.append(f"❌ Email inválido: {aa_user}. Debe ser @bcp.com.pe")
    
    # Validar que exista el archivo de controles
    control_file = f"{base_path}Controles_Parent_domain.xlsx"
    if not os.path.exists(control_file):
        errors.append(f"❌ Archivo de control no encontrado: {control_file}")
    
    # Validar que exista la carpeta
    if not os.path.exists(base_path):
        errors.append(f"❌ Ruta base no existe: {base_path}")
    
    # Validar nombre de carpeta
    if not ab_carpeta or ab_carpeta.strip() == "":
        errors.append("❌ Nombre de carpeta no puede estar vacío")
    
    if errors:
        error_msg = "\n".join(["🚨 ERRORES DE VALIDACIÓN:"] + errors)
        raise ValueError(error_msg)
    
    print("✅ Validación de entradas completada exitosamente")


# Ejecutar validación
try:
    validate_inputs(aa_user, base_path, ab_carpeta)
except ValueError as e:
    print(str(e))
    dbutils.notebook.exit("ERROR: Validación de entradas falló")

# COMMAND ----------

# ========================================
# VALIDACIÓN DE CARPETA (Código Original Mejorado)
# ========================================

# Busca la carpeta ignorando mayúsculas
result = subprocess.run(["ls", base_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
folders_list = result.stdout.splitlines()

match = next((f for f in folders_list if f.lower() == ab_carpeta), None)

if match:
    base_path_carpeta = f"{base_path}{match}/"
    print(f"✅ Carpeta encontrada: {base_path_carpeta}")
else:
    error_msg = f"❌ En la ruta: {base_path} no se encontró la carpeta: {ab_carpeta}"
    print(error_msg)
    raise ValueError(error_msg)

# 🆕 NUEVO: Inicializar logger después de validar carpeta
logger = setup_logger(base_path_carpeta, ab_carpeta)

# COMMAND ----------

# ========================================
# FUNCIONES AUXILIARES (Código Original)
# ========================================

def read_excel_files(
    ruta: str,
    prefijo: str or list = None,
    extension: str = None,
    hoja: str = "Sheet1",
    skip_filas: int = 0,
    agregar_nombre_archivo: bool = False,
    archivo_especifico: str = None
) -> DataFrame:
    """
    Lee uno o varios archivos Excel de una carpeta, según prefijo, extensión y hoja específica.
 
    Args:
        ruta (str): Carpeta donde están los archivos.
        prefijo (str or list, opcional): Prefijo o lista de prefijos que deben tener los archivos.
        extension (str, opcional): Extensión de archivo, como ".xlsx" o ".xlsm".
        hoja (str, opcional): Nombre de la hoja de Excel a leer.
        skip_filas (int, opcional): Número de filas a omitir.
        agregar_nombre_archivo (bool, opcional): Si se agrega columna con nombre de archivo.
        archivo_especifico (str, opcional): Si se indica, se leerá solo ese archivo.
 
    Returns:
        DataFrame: DataFrame de Spark con los datos consolidados.
    """
    # 🆕 MEJORADO: Agregado logging
    logger.info(f"📂 Leyendo archivos de: {ruta}")
    logger.info(f"   Prefijo: {prefijo}, Extensión: {extension}, Hoja: {hoja}")
 
    if archivo_especifico:
        archivos_filtrados = [os.path.join(ruta, archivo_especifico)]
    else:
        if isinstance(prefijo, list):
            prefijo = tuple(prefijo)
 
        archivos = [
            f for f in os.listdir(ruta)
            if (prefijo is None or f.startswith(prefijo))
            and (extension is None or f.endswith(extension))
        ]
        archivos_filtrados = [os.path.join(ruta, f) for f in archivos]
 
    if not archivos_filtrados:
        # 🆕 MEJORADO: Usar logger en lugar de raise directo
        logger.error(f"No se encontraron archivos con prefijo: {prefijo} en {ruta}")
        raise FileNotFoundError(f"No se encontraron archivos con el prefijo:{prefijo} en {ruta}")
    
    logger.info(f"   ✓ {len(archivos_filtrados)} archivo(s) encontrado(s)")
 
    df_list = []
    for archivo in archivos_filtrados:
        try:
            xls = pd.ExcelFile(archivo)
            if hoja in xls.sheet_names:
                df = pd.read_excel(archivo, sheet_name=hoja, skiprows=skip_filas, engine="openpyxl")
                if agregar_nombre_archivo:
                    df["ARCHIVO"] = os.path.basename(archivo)
 
                if not df.empty:
                    df_list.append(df)
                    logger.info(f"   ✓ Leído: {os.path.basename(archivo)} ({len(df)} filas)")
                else:
                    logger.warning(f"   ⚠ Archivo vacío: {archivo}")
            else:
                logger.warning(f"   ⚠ El archivo {archivo} no tiene la hoja '{hoja}'")
        except Exception as e:
            logger.error(f"   ❌ Error al leer {archivo}: {e}")
 
    if df_list:
        df_consolidado_pd = pd.concat(df_list, ignore_index=True).dropna(how="all")
        df_consolidado_pd.columns = [col.upper() for col in df_consolidado_pd.columns]
        logger.info(f"✅ Consolidado exitoso: {len(df_consolidado_pd)} filas totales")
        return spark.createDataFrame(df_consolidado_pd)
    else:
        logger.error("No se pudieron consolidar los archivos")
        raise ValueError("No se pudieron consolidar los archivos correctamente.")


def remove_accents_df(df: DataFrame) -> DataFrame:
    """Elimina tildes de los nombres de columnas del DataFrame"""
    def remove_accents(text: str) -> str:
        return ''.join(c for c in unicodedata.normalize('NFD', text) if not unicodedata.combining(c))
    
    new_columns = [remove_accents(col) for col in df.columns]
    logger.debug(f"Columnas sin tildes: {new_columns[:5]}...")  # Mostrar solo primeras 5
    return df.toDF(*new_columns)


def remove_accents(text):
    """Elimina tildes del texto"""
    if text is None:
        return ""
    return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')


def read_fields(df, capa, validar=None):
    """
    Lee los registros del archivo Controles_Parent_domain y los carga como lista.
 
    Args:
        df: DataFrame del control_parent_domain.
        capa: Capa a validar.
        validar: Filtro adicional opcional.
        
    Returns:
        tuple: (list_fields, nom_prefijos, nom_hojas)
    """
    # 🆕 MEJORADO: Mejor logging
    logger.info(f"🔍 Extrayendo campos para capa: {capa}")
    
    df = df.withColumn("ORDEN", monotonically_increasing_id())
    
    df_filter = df.filter(upper(col("CAPA")) == capa.upper())
    if validar is not None:
        df_filter = df_filter.filter(upper(col("VALIDAR")) == validar.upper())
 
    def extract_ordered_unique_values(df_interno, col_name):
        window_spec = Window.partitionBy(col_name).orderBy("ORDEN")
        df_ranked = df_interno.withColumn("rn", row_number().over(window_spec))
        result_rows = (
            df_ranked.filter(col("rn") == 1)
            .select(col_name, "ORDEN")
            .where(col(col_name).isNotNull())
            .orderBy("ORDEN")
            .collect()
        )
        return [remove_accents(row[col_name].upper()) for row in result_rows]
 
    # NOM_CAMPO_ARCHIVO
    if "NOM_CAMPO_ARCHIVO" in df_filter.columns:
        list_fields = extract_ordered_unique_values(df_filter, "NOM_CAMPO_ARCHIVO")
    else:
        raise ValueError("No se encontró el campo: NOM_CAMPO_ARCHIVO")
 
    # NOM_PREFIJO_ARCHIVO
    if "NOM_PREFIJO_ARCHIVO" in df_filter.columns:
        nom_prefijos = extract_ordered_unique_values(df_filter, "NOM_PREFIJO_ARCHIVO")
    else:
        raise ValueError("No se encontró el campo: NOM_PREFIJO_ARCHIVO")
 
    # NOM_HOJA
    if "NOM_HOJA" in df_filter.columns:
        nom_hoja = extract_ordered_unique_values(df_filter, "NOM_HOJA")
    else:
        raise ValueError("No se encontró el campo: NOM_HOJA")
    
    logger.info(f"   ✓ {len(list_fields)} campos extraídos para {capa}")
 
    return list_fields, nom_prefijos, nom_hoja


def validate_columns(df: DataFrame, required_columns: List[str], df_name: str) -> None:
    """Valida que existan columnas requeridas en el DataFrame"""
    missing = set(required_columns) - set(df.columns)
    if missing:
        logger.error(f"Columnas faltantes en '{df_name}': {missing}")
        raise ValueError(f"El archivo '{df_name}' no contiene las siguientes columnas necesarias: {missing}")
    logger.debug(f"✓ Todas las columnas requeridas presentes en '{df_name}'")


def clear_column(df, c):
    """Limpia valores nulos y caracteres especiales de una columna"""
    return df.withColumn(c, when(trim(lower(col(c))).isin("nan", ".."), "").otherwise(trim(col(c))))


def normalize_and_names(df: DataFrame, to_upper_columns: list) -> DataFrame:
    """Normaliza nombres de columnas y convierte valores a mayúsculas"""
    new_columns_names = [c.strip().replace(" ","_") for c in df.columns]
    df = df.toDF(*new_columns_names)

    for c in to_upper_columns:
        c_norm = c.strip().replace(" ","_")
        if c_norm in df.columns:
            df = df.withColumn(c_norm, upper(trim(col(c_norm))))
        else:
            logger.warning(f"⚠ Columna '{c}' no existe en el dataframe")
    
    return df


def process_dataframes(da_df: DataFrame, qa_df: DataFrame, tabla: str, campo: str, capa_nombre: str) -> DataFrame:
    """
    Valida campos DA vs QA y genera reporte de diferencias.
    
    🆕 MEJORADO: Agregado parámetro capa_nombre para mejor logging
    """
    columns_order = da_df.columns

    qa_comp = qa_df.selectExpr(
            f"TABLA_FISICA as {tabla}",
            f"CAMPO_FISICO as {campo}",
            "True as FLAG_QA"
        )

    da_df_flag = da_df.selectExpr("*", "True as FLAG_DA")
    
    joined_df = da_df_flag.join(qa_comp, on=[tabla, campo], how="full")
    
    resultado = joined_df.withColumn(
        "QA_Campo",
        when(col("FLAG_DA").isNotNull() & col("FLAG_QA").isNotNull(), "OK")
        .when(col("FLAG_DA").isNotNull() & col("FLAG_QA").isNull(), "No esta modelado")
        .when(col("FLAG_DA").isNull() & col("FLAG_QA").isNotNull(), "En QA pero no en DA")
    ).drop("FLAG_DA","FLAG_QA")

    total = resultado.count()
    faltantes = resultado.filter(col("QA_Campo") != "OK").count()
    
    # 🆕 MEJORADO: Mejor logging con logger
    if faltantes == 0:
        logger.info(f"✅ Todos los {total} campos de {capa_nombre} coinciden con QA")
    else:
        logger.warning(f"⚠️ De {total} campos de {capa_nombre}, {faltantes} no están en QA")
    
    columns_end = [c for c in columns_order if c in resultado.columns] + \
                  [c for c in resultado.columns if c not in columns_order]
    return resultado.select(*columns_end)


def standardize_columns(df, sufijo):
    """Renombra columnas quitando sufijo (DDV o UDV)"""
    columns_rename = []
    for c in df.columns:
        new_name = re.sub(rf"\s*{sufijo}$", "", c.strip(), flags = re.IGNORECASE)
        columns_rename.append(col(c).alias(new_name))
    return df.select(columns_rename)

# COMMAND ----------

# ========================================
# 🆕 NUEVAS FUNCIONES DE MEJORA
# ========================================

def compare_with_previous_report(current_df: pd.DataFrame, base_path_carpeta: str, ab_carpeta: str) -> Optional[pd.DataFrame]:
    """
    Compara el reporte actual con el anterior para detectar cambios.
    
    Args:
        current_df: DataFrame actual con resultados
        base_path_carpeta: Ruta de la carpeta
        ab_carpeta: Nombre de carpeta
        
    Returns:
        DataFrame con cambios detectados o None si no hay reporte anterior
    """
    logger.info("🔄 Buscando reportes anteriores para comparación...")
    
    # Buscar reportes anteriores
    try:
        existing_reports = sorted([
            f for f in os.listdir(base_path_carpeta) 
            if f.startswith(f"Reporte_Validacion_QA_{ab_carpeta}_") and f.endswith(".xlsx")
        ])
    except Exception as e:
        logger.warning(f"No se pudieron listar reportes anteriores: {e}")
        return None
    
    if len(existing_reports) < 1:
        logger.info("   ℹ️ No hay reportes anteriores para comparar")
        return None
    
    previous_report = existing_reports[-1]
    logger.info(f"   📄 Comparando con: {previous_report}")
    
    try:
        # Leer reporte anterior
        previous_df = pd.read_excel(
            f"{base_path_carpeta}{previous_report}", 
            sheet_name="Consolidado_DA"
        )
        
        # Detectar cambios
        current_issues = set(current_df[current_df['QA_Campo'] != 'OK']['CAMPO'].values) if 'CAMPO' in current_df.columns else set()
        previous_issues = set(previous_df[previous_df['QA_Campo'] != 'OK']['CAMPO'].values) if 'CAMPO' in previous_df.columns else set()
        
        new_issues = current_issues - previous_issues
        resolved_issues = previous_issues - current_issues
        
        # Crear DataFrame de cambios
        changes_data = {
            'Tipo': ['🆕 Nuevos Problemas', '✅ Problemas Resueltos', '📊 Problemas Persistentes'],
            'Cantidad': [
                len(new_issues), 
                len(resolved_issues),
                len(current_issues & previous_issues)
            ],
            'Detalle': [
                ', '.join(list(new_issues)[:10]) + ('...' if len(new_issues) > 10 else '') if new_issues else 'Ninguno',
                ', '.join(list(resolved_issues)[:10]) + ('...' if len(resolved_issues) > 10 else '') if resolved_issues else 'Ninguno',
                f"{len(current_issues & previous_issues)} campos"
            ]
        }
        
        changes_df = pd.DataFrame(changes_data)
        
        logger.info(f"   🆕 Nuevos problemas: {len(new_issues)}")
        logger.info(f"   ✅ Problemas resueltos: {len(resolved_issues)}")
        logger.info(f"   📊 Problemas persistentes: {len(current_issues & previous_issues)}")
        
        return changes_df
        
    except Exception as e:
        logger.warning(f"   ⚠️ Error al comparar con reporte anterior: {e}")
        return None


def create_summary_sheet(writer, pd_resultado_da: pd.DataFrame, resultado_parent_domain_pd: pd.DataFrame, 
                        ab_carpeta: str, changes_df: Optional[pd.DataFrame] = None):
    """
    Crea una hoja de resumen ejecutivo con KPIs y estadísticas.
    
    Args:
        writer: ExcelWriter object
        pd_resultado_da: DataFrame de resultados DA
        resultado_parent_domain_pd: DataFrame de resultados Parent Domain
        ab_carpeta: Nombre de carpeta
        changes_df: DataFrame con cambios respecto a reporte anterior (opcional)
    """
    logger.info("📊 Creando hoja de resumen ejecutivo...")
    
    # Calcular métricas DA
    total_campos = len(pd_resultado_da)
    campos_ok = len(pd_resultado_da[pd_resultado_da['QA_Campo'] == 'OK'])
    campos_no_modelados = len(pd_resultado_da[pd_resultado_da['QA_Campo'] == 'No esta modelado'])
    campos_qa_sin_da = len(pd_resultado_da[pd_resultado_da['QA_Campo'] == 'En QA pero no en DA'])
    cobertura_pct = (campos_ok/total_campos*100) if total_campos > 0 else 0
    
    # Métricas Parent Domain
    pd_ok = len(resultado_parent_domain_pd[resultado_parent_domain_pd['Observaciones'] == 'OK'])
    pd_total = len(resultado_parent_domain_pd)
    pd_pct = (pd_ok/pd_total*100) if pd_total > 0 else 0
    
    # Crear secciones del resumen
    summary_sections = []
    
    # Sección 1: Información General
    summary_sections.extend([
        {'Categoría': '📋 INFORMACIÓN GENERAL', 'Métrica': '', 'Valor': ''},
        {'Categoría': '', 'Métrica': 'Fecha de Validación', 'Valor': datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {'Categoría': '', 'Métrica': 'Carpeta Analizada', 'Valor': ab_carpeta},
        {'Categoría': '', 'Métrica': 'Usuario', 'Valor': aa_user},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Sección 2: Métricas de Campos
    summary_sections.extend([
        {'Categoría': '🎯 VALIDACIÓN DE CAMPOS', 'Métrica': '', 'Valor': ''},
        {'Categoría': '', 'Métrica': 'Total Campos Validados', 'Valor': total_campos},
        {'Categoría': '', 'Métrica': '✅ Campos OK', 'Valor': campos_ok},
        {'Categoría': '', 'Métrica': '⚠️ Campos No Modelados', 'Valor': campos_no_modelados},
        {'Categoría': '', 'Métrica': '🔴 Campos en QA sin DA', 'Valor': campos_qa_sin_da},
        {'Categoría': '', 'Métrica': '📊 % Cobertura', 'Valor': f"{cobertura_pct:.2f}%"},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Sección 3: Parent Domain
    summary_sections.extend([
        {'Categoría': '🏛️ PARENT DOMAIN', 'Métrica': '', 'Valor': ''},
        {'Categoría': '', 'Métrica': 'Total Validados', 'Valor': pd_total},
        {'Categoría': '', 'Métrica': '✅ Parent Domain OK', 'Valor': pd_ok},
        {'Categoría': '', 'Métrica': '⚠️ Parent Domain con Observaciones', 'Valor': pd_total - pd_ok},
        {'Categoría': '', 'Métrica': '📊 % Cumplimiento', 'Valor': f"{pd_pct:.2f}%"},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Sección 4: Comparación con reporte anterior (si existe)
    if changes_df is not None:
        summary_sections.extend([
            {'Categoría': '🔄 COMPARACIÓN CON REPORTE ANTERIOR', 'Métrica': '', 'Valor': ''},
        ])
        for _, row in changes_df.iterrows():
            summary_sections.append({
                'Categoría': '', 
                'Métrica': row['Tipo'], 
                'Valor': row['Cantidad']
            })
        summary_sections.append({'Categoría': '', 'Métrica': '', 'Valor': ''})
    
    # Sección 5: Recomendaciones
    recommendations = []
    if campos_no_modelados > 0:
        recommendations.append(f"⚠️ {campos_no_modelados} campos requieren modelado en QA")
    if campos_qa_sin_da > 0:
        recommendations.append(f"🔴 {campos_qa_sin_da} campos en QA no están documentados en DA")
    if pd_total - pd_ok > 0:
        recommendations.append(f"⚠️ {pd_total - pd_ok} campos con Parent Domain incorrecto")
    
    if recommendations:
        summary_sections.extend([
            {'Categoría': '💡 RECOMENDACIONES', 'Métrica': '', 'Valor': ''},
        ])
        for i, rec in enumerate(recommendations, 1):
            summary_sections.append({
                'Categoría': '', 
                'Métrica': f'{i}. {rec}', 
                'Valor': ''
            })
    
    # Crear DataFrame
    summary_df = pd.DataFrame(summary_sections)
    
    # Escribir a Excel
    summary_df.to_excel(writer, sheet_name="📊 Resumen Ejecutivo", index=False)
    
    # Formatear hoja
    ws = writer.sheets["📊 Resumen Ejecutivo"]
    
    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    category_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    bold_font = Font(bold=True, size=10)
    
    # Formatear encabezados
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Formatear contenido
    for row in range(2, len(summary_df) + 2):
        categoria_cell = ws.cell(row=row, column=1)
        metrica_cell = ws.cell(row=row, column=2)
        valor_cell = ws.cell(row=row, column=3)
        
        # Categorías principales (encabezados de sección)
        if categoria_cell.value and categoria_cell.value.startswith(('📋', '🎯', '🏛️', '🔄', '💡')):
            categoria_cell.fill = category_fill
            categoria_cell.font = white_font
            ws.merge_cells(f'A{row}:C{row}')
            categoria_cell.alignment = Alignment(horizontal='left', vertical='center')
        # Métricas
        elif metrica_cell.value:
            if row % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = metric_fill
            
            # Negrita para métricas importantes
            if any(x in str(metrica_cell.value) for x in ['Total', 'Cobertura', 'Cumplimiento']):
                metrica_cell.font = bold_font
                valor_cell.font = bold_font
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    
    logger.info("   ✅ Hoja de resumen creada exitosamente")


def add_excel_features(workbook, sheet_name: str, num_cols: int, num_rows: int):
    """
    Agrega características avanzadas a las hojas de Excel.
    
    Args:
        workbook: Objeto workbook de openpyxl
        sheet_name: Nombre de la hoja
        num_cols: Número de columnas
        num_rows: Número de filas
    """
    ws = workbook[sheet_name]
    
    # Agregar autofiltro
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
    
    # Congelar primera fila
    ws.freeze_panes = "A2"
    
    logger.debug(f"   ✓ Filtros y congelación aplicados a '{sheet_name}'")


def manage_report_versions(base_path_carpeta: str, ab_carpeta: str, max_versions: int = 10):
    """
    Mantiene solo las últimas N versiones de reportes.
    
    Args:
        base_path_carpeta: Ruta de la carpeta
        ab_carpeta: Nombre de carpeta
        max_versions: Número máximo de versiones a mantener
    """
    logger.info(f"🗂️ Gestionando versiones de reportes (max: {max_versions})...")
    
    try:
        reports = sorted([
            f for f in os.listdir(base_path_carpeta)
            if f.startswith(f"Reporte_Validacion_QA_{ab_carpeta}_") and f.endswith(".xlsx")
        ])
        
        if len(reports) > max_versions:
            # Crear carpeta de archivo
            archive_path = f"{base_path_carpeta}archivo/"
            os.makedirs(archive_path, exist_ok=True)
            
            # Mover reportes antiguos
            for old_report in reports[:-max_versions]:
                shutil.move(
                    f"{base_path_carpeta}{old_report}",
                    f"{archive_path}{old_report}"
                )
            
            archived_count = len(reports) - max_versions
            logger.info(f"   📦 {archived_count} reporte(s) archivado(s)")
        else:
            logger.info(f"   ✓ {len(reports)} reporte(s) actual(es), no se requiere archivado")
            
    except Exception as e:
        logger.warning(f"   ⚠️ Error al gestionar versiones: {e}")

# COMMAND ----------

# ========================================
# LECTURA DE PARÁMETROS DE VALIDACIÓN
# ========================================

logger.info("="*80)
logger.info("📖 FASE 1: Lectura de Parámetros de Validación")
logger.info("="*80)

# Leer archivo de controles
df_control = read_excel_files(
    ruta=base_path,
    prefijo="Controles_Parent_domain",
    extension=".xlsx",
    hoja="parametros",
    skip_filas=0,
    agregar_nombre_archivo=False
)

# 🆕 MEJORADO: Cachear DataFrame que se usa múltiples veces
df_control.cache()

# Obtener parámetros DDV
fields_ddv, prefijo_ddv, hoja_ddv = read_fields(df_control, "DDV")
fields_ddv_validation, _, _ = read_fields(df_control, "DDV", "SI")

logger.info(f"DDV - Campos: {len(fields_ddv)}, Prefijo: {prefijo_ddv}, Hoja: {hoja_ddv}")
logger.info(f"DDV - Campos validación: {len(fields_ddv_validation)}")

# Obtener parámetros UDV
fields_udv, prefijo_udv, hoja_udv = read_fields(df_control, "UDV")
fields_udv_validation, _, _ = read_fields(df_control, "UDV", "SI")

logger.info(f"UDV - Campos: {len(fields_udv)}, Prefijo: {prefijo_udv}, Hoja: {hoja_udv}")

# Obtener parámetros QA
fields_qa, prefijo_qa, hoja_qa = read_fields(df_control, "QA")

logger.info(f"QA - Campos: {len(fields_qa)}, Prefijo: {prefijo_qa}, Hoja: {hoja_qa}")

# Liberar cache
df_control.unpersist()

# COMMAND ----------

# ========================================
# LISTADO DE ARCHIVOS EN WORKSPACE
# ========================================

logger.info("="*80)
logger.info("📂 FASE 2: Listado de Archivos en Workspace")
logger.info("="*80)

output = subprocess.run(["ls", base_path_carpeta], capture_output=True, text=True)
file_list = output.stdout.split("\n")

# Filtrar archivos por tipo
qa_file_list = [base_path_carpeta + f for f in file_list if f.startswith(prefijo_qa[0]) and f.endswith('.xlsx')]
da_ddv_file_list = [base_path_carpeta + f for f in file_list if f.startswith(prefijo_ddv[0]) and f.endswith('.xlsm')]
da_udv_file_list = [base_path_carpeta + f for f in file_list if f.startswith(prefijo_udv[0]) and f.endswith('.xlsm')]

# Nombres de archivos
qa_file_names = [os.path.basename(f) for f in qa_file_list]
da_ddv_file_names = [os.path.basename(f) for f in da_ddv_file_list]
da_udv_file_names = [os.path.basename(f) for f in da_udv_file_list]

logger.info(f"Archivos QA encontrados: {len(qa_file_names)}")
for f in qa_file_names:
    logger.info(f"   - {f}")

logger.info(f"Archivos DA DDV encontrados: {len(da_ddv_file_names)}")
for f in da_ddv_file_names:
    logger.info(f"   - {f}")

logger.info(f"Archivos DA UDV encontrados: {len(da_udv_file_names)}")
for f in da_udv_file_names:
    logger.info(f"   - {f}")

# 🆕 NUEVO: Validar que haya al menos un archivo QA
if not qa_file_list:
    logger.error("❌ No se encontraron archivos QA_REPORTE_*.xlsx")
    raise FileNotFoundError("No se encontraron archivos QA_REPORTE_*.xlsx en la carpeta")

if not da_ddv_file_list and not da_udv_file_list:
    logger.error("❌ No se encontraron archivos de Documento de Alcance")
    raise FileNotFoundError("No se encontraron archivos DA en la carpeta")

# COMMAND ----------

# ========================================
# CONSOLIDADO DE ARCHIVO QA_REPORTE
# ========================================

logger.info("="*80)
logger.info("🔍 FASE 3: Procesamiento de Archivo QA_REPORTE")
logger.info("="*80)

# 1. Leer archivo
df_qa_consolidado = read_excel_files(
    ruta=base_path_carpeta,
    prefijo=prefijo_qa[0],
    extension=".xlsx",
    hoja=hoja_qa[0],
    skip_filas=0,
    agregar_nombre_archivo=True
)

# 2. Validar columnas
validate_columns(df_qa_consolidado, fields_qa, "QA_REPORTE")

# 3. Limpiar valores nulos
columns_to_clean = df_qa_consolidado.select(fields_qa + ["ARCHIVO"])
df_qa_consolidado_cleaned = reduce(clear_column, columns_to_clean.columns, columns_to_clean)

# 4. Normalizar columnas
column_qa_norm = df_qa_consolidado_cleaned.drop("ARCHIVO", "PARENT_DOMAIN")
df_qa_consolidado_norm = normalize_and_names(df_qa_consolidado_cleaned, column_qa_norm.columns)

# 🆕 MEJORADO: Cachear DataFrame QA que se usa en múltiples joins
df_qa_consolidado_norm.cache()

logger.info(f"✅ QA consolidado: {df_qa_consolidado_norm.count()} registros")

# COMMAND ----------

# ========================================
# PROCESAMIENTO POR CADA CAPA (DDV/UDV)
# ========================================

logger.info("="*80)
logger.info("🏗️ FASE 4: Procesamiento de Capas DDV/UDV")
logger.info("="*80)

capas = [
    {
        "nombre": "DDV",
        "prefijo": prefijo_ddv[0],
        "hoja": hoja_ddv[0],
        "campos": fields_ddv,
        "validacion": fields_ddv_validation
    },
    {
        "nombre": "UDV",
        "prefijo": prefijo_udv[0],
        "hoja": hoja_udv[0],
        "campos": fields_udv,
        "validacion": fields_udv_validation
    }
]

for capa in capas:
    logger.info(f"\n{'='*60}")
    logger.info(f"Procesando Capa: {capa['nombre']}")
    logger.info(f"{'='*60}")
    
    archivos_capa = [f for f in os.listdir(base_path_carpeta) if f.startswith(capa["prefijo"]) and f.endswith(".xlsm")]
    
    if not archivos_capa:
        logger.warning(f"⚠️ No se encontraron archivos para capa {capa['nombre']}")
        continue
    
    try:
        # Leer archivos
        df_consolidado = read_excel_files(
            ruta=base_path_carpeta,
            prefijo=capa["prefijo"],
            extension=".xlsm",
            hoja=capa["hoja"],
            skip_filas=5,
            agregar_nombre_archivo=True
        )
        
        # Agregar orden
        campo_col = f'"CAMPO {capa["nombre"]}"'
        df_consolidado = df_consolidado.selectExpr("*", f"row_number() over (ORDER BY {campo_col}) as ORDEN")
        
        # Limpiar tildes
        df_consolidado = remove_accents_df(df_consolidado)
        
        # Validar columnas
        validate_columns(df_consolidado, capa["campos"], f"Documento de Alcance {capa['nombre']}")
        
        # Limpiar nulos
        columnas_con_archivo = df_consolidado.select(*capa["campos"] + ["ARCHIVO", "ORDEN"])
        df_limpio = reduce(clear_column, columnas_con_archivo.columns, columnas_con_archivo)
        
        # Filtrar filas vacías
        condition = reduce(lambda a, b: a | b, [(col(c).isNotNull()) & (col(c) != "") for c in capa["campos"]])
        df_filtrado = df_limpio.filter(condition)
        
        # Normalizar
        df_normalizado = normalize_and_names(df_filtrado, capa["validacion"])
        
        # Comparar contra QA
        tabla_col = f"TABLA_{capa['nombre']}"
        campo_col = f"CAMPO_{capa['nombre']}"
        
        resultado_df = process_dataframes(
            df_normalizado, 
            df_qa_consolidado_norm, 
            tabla_col, 
            campo_col,
            capa['nombre']  # 🆕 NUEVO: Pasar nombre de capa
        )
        
        # Ordenar columnas
        columns = [col for col in resultado_df.columns if col not in ["ARCHIVO", "ORDEN"]]
        
        if capa["nombre"] == "DDV":
            df_ddv_resultado = resultado_df.select(["ORDEN", "ARCHIVO"] + columns)
        else:
            df_udv_resultado = resultado_df.select(["ORDEN", "ARCHIVO"] + columns)
        
        logger.info(f"✅ Capa {capa['nombre']} procesada exitosamente")
        
    except FileNotFoundError as e:
        logger.error(f"❌ Error en lectura de archivos para {capa['nombre']}: {e}")
    except Exception as e:
        logger.error(f"❌ Error inesperado en {capa['nombre']}: {e}")
        raise

# COMMAND ----------

# ========================================
# CONSOLIDACIÓN DE RESULTADOS DA
# ========================================

logger.info("="*80)
logger.info("🔗 FASE 5: Consolidación de Resultados DA")
logger.info("="*80)

# Validar existencia de DataFrames
try:
    df_ddv_resultado
    logger.info("✓ DataFrame DDV disponible")
except NameError:
    df_ddv_resultado = None
    logger.warning("⚠ DataFrame DDV no disponible")

try:
    df_udv_resultado
    logger.info("✓ DataFrame UDV disponible")
except NameError:
    df_udv_resultado = None
    logger.warning("⚠ DataFrame UDV no disponible")

# Estandarizar columnas
if df_ddv_resultado is not None:
    df_ddv_resultado = standardize_columns(df_ddv_resultado, "DDV")
    logger.info("✓ Columnas DDV estandarizadas")

if df_udv_resultado is not None:
    df_udv_resultado = standardize_columns(df_udv_resultado, "UDV")
    logger.info("✓ Columnas UDV estandarizadas")

# Unión condicional
if df_ddv_resultado is not None and df_udv_resultado is not None:
    df_resultado_da = df_ddv_resultado.unionByName(df_udv_resultado)
    logger.info("✅ DataFrames DDV y UDV unidos exitosamente")
elif df_ddv_resultado is not None:
    df_resultado_da = df_ddv_resultado
    logger.info("✅ Solo DataFrame DDV disponible")
elif df_udv_resultado is not None:
    df_resultado_da = df_udv_resultado
    logger.info("✅ Solo DataFrame UDV disponible")
else:
    logger.error("❌ No se encontraron resultados de ninguna capa")
    raise ValueError("No se encontraron resultados de DDV ni UDV")

logger.info(f"Total registros consolidados: {df_resultado_da.count()}")

# COMMAND ----------

# ========================================
# VALIDACIÓN DE PARENT DOMAIN
# ========================================

logger.info("="*80)
logger.info("🏛️ FASE 6: Validación de Parent Domain")
logger.info("="*80)

# Leer controles Parent Domain
df_control_parent_domain = read_excel_files(
    ruta=base_path,
    prefijo="Controles_Parent_domain",
    extension=".xlsx",
    hoja="controles",
    skip_filas=0,
    agregar_nombre_archivo=False
)

# Limpiar duplicados y nulos
df_control_parent_domain_unique = df_control_parent_domain.dropDuplicates()
df_control_parent_domain_norm = reduce(clear_column, df_control_parent_domain_unique.columns, df_control_parent_domain_unique)

# 🆕 MEJORADO: Cachear DataFrame de controles
df_control_parent_domain_norm.cache()

logger.info(f"Controles Parent Domain: {df_control_parent_domain_norm.count()} registros")

# Lineamiento por tipo
comparacion_lineamiento_df = df_qa_consolidado_norm.join(
    df_control_parent_domain_norm.filter(col("LINEAMIENTO") == "Fijo"),
    on=col("CAMPO_FISICO") == col("CONTROL"),
    how="left"
).select(
    df_qa_consolidado_norm["*"],
    when(col("LINEAMIENTO").isNotNull(), col("LINEAMIENTO")).otherwise("PreFijo").alias("LINEAMIENTO_PD")
)

# Separar Fijos y Prefijos
df_fijo = comparacion_lineamiento_df.filter(col("LINEAMIENTO_PD") == "Fijo")
df_prefijo = comparacion_lineamiento_df.filter(col("LINEAMIENTO_PD") == "PreFijo")

logger.info(f"Campos Fijos: {df_fijo.count()}")
logger.info(f"Campos PreFijo: {df_prefijo.count()}")

# Procesamiento Fijos
if df_fijo.count() > 0:
    comparacion_control_fijo_df = df_fijo.join(
        df_control_parent_domain_norm.select("CONTROL", "LINEAMIENTO"),
        (df_fijo.CAMPO_FISICO == df_control_parent_domain_norm.CONTROL) &
        (df_fijo.LINEAMIENTO_PD == df_control_parent_domain_norm.LINEAMIENTO),
        "left"
    ).select(
        df_fijo["*"],
        df_control_parent_domain_norm.CONTROL.alias("CONTROL_PR")
    )
    
    df_comparacion_parent_domain_fijo = comparacion_control_fijo_df.join(
        df_control_parent_domain_norm.filter(col("LINEAMIENTO") == "Fijo"),
        (col("CONTROL_PR") == col("CONTROL")),
        "left"
    ).select(
        comparacion_control_fijo_df["*"],
        when(comparacion_control_fijo_df.PARENT_DOMAIN == df_control_parent_domain_norm.PARENT_DOMAIN, "OK")
            .otherwise(col("WARNING_PARENT_DOMAIN")).alias("Observaciones")
    ).drop("LINEAMIENTO_PD", "CONTROL_PR")
    
    logger.info(f"✓ Procesamiento Fijos completado")

# Procesamiento Prefijos
if df_prefijo.count() > 0:
    # Lista de controles PreFijo
    prefijo_fields_len = (
        df_control_parent_domain_norm
        .filter(col("LINEAMIENTO") == "PreFijo")
        .select(
            col("CONTROL"),
            length(trim(col("CONTROL"))).cast("int").alias("LEN_CONTROL"),
            col("PARENT_DOMAIN")
        )
        .dropDuplicates()
        .collect()
    )
    
    prefijo_fields_len = sorted(prefijo_fields_len, key=lambda row: row["LEN_CONTROL"], reverse=True)
    
    # Crear columna CONTROL_PR
    expr = F.lit("OBSERVADO")
    for row in reversed(prefijo_fields_len):
        prefix = row["CONTROL"].upper()
        length_prefix = row["LEN_CONTROL"]
        
        expr = when(
            substring(trim(upper(col("CAMPO_FISICO"))), 1, length_prefix) == prefix,
            prefix
        ).otherwise(expr)
    
    comparacion_control_prefijo_df = df_prefijo.select("*", expr.alias("CONTROL_PR"))
    
    df_comparacion_parent_domain_prefijo = comparacion_control_prefijo_df.join(
        df_control_parent_domain_norm.filter(col("LINEAMIENTO") == "PreFijo"),
        (col("CONTROL_PR") == col("CONTROL")),
        "left"
    ).select(
        comparacion_control_prefijo_df["*"],
        when(comparacion_control_prefijo_df.PARENT_DOMAIN == df_control_parent_domain_norm.PARENT_DOMAIN, "OK")
            .otherwise(col("WARNING_PARENT_DOMAIN")).alias("Observaciones")
    ).drop("LINEAMIENTO_PD", "CONTROL_PR")
    
    logger.info(f"✓ Procesamiento Prefijos completado")

# Unir resultados
if df_fijo.count() > 0 and df_prefijo.count() > 0:
    comparacion_control_df = df_comparacion_parent_domain_fijo.unionByName(df_comparacion_parent_domain_prefijo)
elif df_fijo.count() > 0:
    comparacion_control_df = df_comparacion_parent_domain_fijo
elif df_prefijo.count() > 0:
    comparacion_control_df = df_comparacion_parent_domain_prefijo
else:
    comparacion_control_df = None
    logger.warning("⚠️ No hay datos de comparación Parent Domain")

# Validación final Parent Domain
if comparacion_control_df is not None:
    columns = [col for col in comparacion_control_df.columns if col != "ARCHIVO"]
    df_comparacion_parent_domain = comparacion_control_df.select(["ARCHIVO"] + columns)
    df_comparacion_parent_domain = df_comparacion_parent_domain \
        .filter(col("ORDEN_FISICO").isNotNull()) \
        .selectExpr("*", "cast(ORDEN_FISICO as int) as ORDEN_FISICO_2") \
        .orderBy("TABLA_FISICA", col("ORDEN_FISICO").cast("int"))
    
    logger.info(f"✅ Validación Parent Domain completada: {df_comparacion_parent_domain.count()} registros")

# Liberar caches
df_qa_consolidado_norm.unpersist()
df_control_parent_domain_norm.unpersist()

# COMMAND ----------

# ========================================
# EXPORTACIÓN DE RESULTADOS
# ========================================

logger.info("="*80)
logger.info("💾 FASE 7: Exportación de Resultados")
logger.info("="*80)

# 🆕 NUEVO: Verificar modo dry-run
if dry_run:
    logger.info("🧪 MODO DRY-RUN ACTIVADO: No se exportarán archivos")
    logger.info("Vista previa de resultados:")
    # EXCLUDE_ANALYZER: DataFrame.toPandas
    display(df_resultado_da.limit(10).toPandas())
    # EXCLUDE_ANALYZER: DataFrame.toPandas
    display(df_comparacion_parent_domain.limit(10).toPandas())
    dbutils.notebook.exit("Proceso completado en modo DRY-RUN")

# Timestamp para el archivo
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

# Convertir DataFrames a Pandas
logger.info("Convirtiendo DataFrames de Spark a Pandas...")
# EXCLUDE_ANALYZER: DataFrame.toPandas
pd_resultado_da = df_resultado_da.toPandas()
pd_resultado_da["ORDEN"] = pd.to_numeric(pd_resultado_da["ORDEN"], errors="coerce")
pd_resultado_da = pd_resultado_da.sort_values(by="ORDEN", ascending=True)

# EXCLUDE_ANALYZER: DataFrame.toPandas
resultado_parent_domain_pd = df_comparacion_parent_domain.toPandas()
resultado_parent_domain_pd = resultado_parent_domain_pd.drop("ORDEN_FISICO_2", axis=1)

logger.info(f"✓ DataFrame DA: {len(pd_resultado_da)} filas")
logger.info(f"✓ DataFrame Parent Domain: {len(resultado_parent_domain_pd)} filas")

# 🆕 NUEVO: Comparar con reporte anterior
changes_df = compare_with_previous_report(pd_resultado_da, base_path_carpeta, ab_carpeta)

# Ruta de exportación
path_export = f"{base_path_carpeta}Reporte_Validacion_QA_{ab_carpeta}_{timestamp}.xlsx"

# 🆕 MEJORADO: Estilos mejorados
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
zebra_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB34", end_color="FFEB34", fill_type="solid")
alert_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

logger.info("Creando archivo Excel con múltiples hojas...")

# Crear Excel con múltiples hojas
with pd.ExcelWriter(path_export, engine="openpyxl") as writer:
    
    # 🆕 NUEVO: Hoja de Resumen Ejecutivo
    create_summary_sheet(writer, pd_resultado_da, resultado_parent_domain_pd, ab_carpeta, changes_df)
    
    # 🆕 NUEVO: Hoja de Comparación (si existe)
    if changes_df is not None:
        changes_df.to_excel(writer, sheet_name="🔄 Comparación", index=False)
        ws_changes = writer.sheets["🔄 Comparación"]
        
        # Formatear encabezados
        for col in range(1, len(changes_df.columns) + 1):
            cell = ws_changes.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = white_font
        
        # Ajustar anchos
        ws_changes.column_dimensions['A'].width = 30
        ws_changes.column_dimensions['B'].width = 15
        ws_changes.column_dimensions['C'].width = 80
        
        logger.info("   ✓ Hoja de Comparación creada")
    
    # Hoja Consolidado_DA
    pd_resultado_da.to_excel(writer, sheet_name="Consolidado_DA", index=False)
    ws1 = writer.sheets["Consolidado_DA"]
    
    num_cols_1 = len(pd_resultado_da.columns)
    num_rows_1 = len(pd_resultado_da)
    estado_col_index_1 = pd_resultado_da.columns.get_loc("QA_Campo") + 1
    
    # Formatear encabezados
    for col in range(1, num_cols_1 + 1):
        cell = ws1.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Formatear filas
    for row in range(2, num_rows_1 + 2):
        estado_cell = ws1.cell(row=row, column=estado_col_index_1)
        apply_yellow = estado_cell.value != "OK"
        apply_alert = estado_cell.value == "En QA pero no en DA"
        
        for col in range(1, num_cols_1 + 1):
            cell = ws1.cell(row=row, column=col)
            cell.border = border
            
            if apply_alert:
                cell.fill = alert_fill
            elif apply_yellow:
                cell.fill = yellow_fill
            elif row % 2 == 0:
                cell.fill = zebra_fill
    
    # Ajustar anchos
    for col_idx, column in enumerate(pd_resultado_da.columns, start=1):
        max_length = max(pd_resultado_da[column].astype(str).map(len).max(), len(column))
        ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    logger.info("   ✓ Hoja Consolidado_DA creada")
    
    # Hoja QA_MODELO
    resultado_parent_domain_pd.to_excel(writer, sheet_name="QA_MODELO", index=False)
    ws2 = writer.sheets["QA_MODELO"]
    
    num_cols_2 = len(resultado_parent_domain_pd.columns)
    num_rows_2 = len(resultado_parent_domain_pd)
    estado_col_index_2 = resultado_parent_domain_pd.columns.get_loc("Observaciones") + 1
    
    # Formatear encabezados
    for col in range(1, num_cols_2 + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Formatear filas
    for row in range(2, num_rows_2 + 2):
        estado_cell = ws2.cell(row=row, column=estado_col_index_2)
        apply_yellow = estado_cell.value != "OK"
        
        for col in range(1, num_cols_2 + 1):
            cell = ws2.cell(row=row, column=col)
            cell.border = border
            
            if apply_yellow:
                cell.fill = yellow_fill
            elif row % 2 == 0:
                cell.fill = zebra_fill
    
    # Ajustar anchos
    for col_idx, column in enumerate(resultado_parent_domain_pd.columns, start=1):
        max_length = max(resultado_parent_domain_pd[column].astype(str).apply(len).max(), len(column))
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    logger.info("   ✓ Hoja QA_MODELO creada")
    
    # 🆕 NUEVO: Agregar características avanzadas de Excel
    wb = writer.book
    add_excel_features(wb, "Consolidado_DA", num_cols_1, num_rows_1)
    add_excel_features(wb, "QA_MODELO", num_cols_2, num_rows_2)

logger.info(f"✅ Reporte exportado exitosamente: {path_export}")

# 🆕 NUEVO: Gestionar versiones de reportes
manage_report_versions(base_path_carpeta, ab_carpeta, max_versiones)

# COMMAND ----------

# ========================================
# 🆕 NUEVO: RESUMEN FINAL Y CIERRE
# ========================================

logger.info("="*80)
logger.info("🎉 PROCESO COMPLETADO EXITOSAMENTE")
logger.info("="*80)

# Estadísticas finales
total_campos = len(pd_resultado_da)
campos_ok = len(pd_resultado_da[pd_resultado_da['QA_Campo'] == 'OK'])
campos_problemas = total_campos - campos_ok
cobertura = (campos_ok / total_campos * 100) if total_campos > 0 else 0

logger.info(f"""
📊 ESTADÍSTICAS FINALES:
   • Total Campos Validados: {total_campos}
   • Campos OK: {campos_ok} ({cobertura:.2f}%)
   • Campos con Problemas: {campos_problemas}
   • Archivo Generado: {os.path.basename(path_export)}
   • Ubicación: {base_path_carpeta}
""")

logger.info("="*80)
logger.info("✅ Fin del proceso")
logger.info("="*80)

# Mostrar mensaje final
print("\n" + "="*80)
print("🎉 PROCESO DE VALIDACIÓN QA COMPLETADO")
print("="*80)
print(f"📄 Reporte generado: {os.path.basename(path_export)}")
print(f"📊 Cobertura: {cobertura:.2f}%")
print(f"✅ Campos OK: {campos_ok}/{total_campos}")
if campos_problemas > 0:
    print(f"⚠️ Campos con problemas: {campos_problemas}")
print("="*80)